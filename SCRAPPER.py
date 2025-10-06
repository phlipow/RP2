from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import spacy
from collections import Counter

# Caminho do Excel
excel_path = r"C:\Users\ATAF IP\COMPRAS\WEBSCRAPPING.xlsx"

# Abre o Excel
wb = load_workbook(excel_path)
ws = wb.active

# Configurações do Chrome (headless para ficar mais rápido)
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)

# Carrega modelo spaCy em português (faça isso fora do loop se possível)
nlp = spacy.load("pt_core_news_sm")

# Loop nas linhas da planilha (a partir da linha 2)
for row in range(2, ws.max_row + 1):
    url = ws[f"A{row}"].value
    if not url:
        continue

    print(f"[Linha {row}] Acessando: {url}")
    try:
        driver.get(url)
        time.sleep(3)  # esperar renderização JS
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # ---------------- Inventores - Resolvido - COLUNA B ----------------
        inventors = []

        try:
            # Seleciona o bloco que contém "important-people"
            important_people = soup.select('dl.important-people.style-scope.patent-result')

            if important_people:
                dl = important_people[0]
                # encontra o <dt> com texto "Inventor" ou "Inventors"
                dt_inventor = dl.find('dt', string=lambda x: x and 'Inventor' in x)
                if dt_inventor:
                    # percorre todos os <dd> depois do <dt> de inventores
                    for dd in dt_inventor.find_all_next('dd'):
                        state = dd.find('state-modifier')  # pega o elemento filho que contém data-inventor
                        if state and state.has_attr('data-inventor'):
                            inventors.append(state['data-inventor'])
                        else:
                            break  # para ao encontrar outro tipo de <dd>

        except Exception as e:
            print(f"  Erro ao buscar inventores: {e}")

        # Salva número de inventores no Excel (coluna B)
        num_inventors = len(inventors)
        ws[f"B{row}"] = num_inventors

        # ---------------- Imagens - Resolvido - COLUNA C ----------------
        image_tags = soup.select('section#thumbnails img')
        ws[f"C{row}"] = len(image_tags)

        # ---------------- Patentes citadas (backward citations)  - Resolvido - COLUNA D ----------------
        backward = []
        citations_section = soup.find('h3', id="patentCitations")
        if citations_section:
            table = citations_section.find_next('div', class_='responsive-table')
            if table:
                modifiers = table.find_all('state-modifier')
                for m in modifiers:
                    data = m.get('data-result', '')
                    if 'patent/' in data:
                        backward.append(data.split('patent/')[-1].split('/')[0])
        ws[f"D{row}"] = len(backward)

        # ---------------- Cited by (forward citations) - Resolvido - COLUNA E ----------------
        cited_by = []
        cited_by_section = soup.find('h3', id="citedBy")
        if cited_by_section:
            table = cited_by_section.find_next('div', class_='responsive-table')
            if table:
                modifiers = table.find_all('state-modifier')
                for m in modifiers:
                    data = m.get('data-result', '')
                    if 'patent/' in data:
                        cited_by.append(data.split('patent/')[-1].split('/')[0])
        ws[f"E{row}"] = len(cited_by)

        # ---------------- Métrica 6: Número de substantivos no título - RESOLVIDO - COLUNA F ---------------

        # Extrai o título da patente
        title_tag = soup.find('h1', class_='scroll-target style-scope patent-result')
        title_text = title_tag.get_text(strip=True) if title_tag else ""

        # Processa o texto com spaCy
        doc = nlp(title_text)

        # Conta o número de substantivos
        num_nouns = sum(1 for token in doc if token.pos_ == "NOUN")

        # Salva no Excel (coluna J, por exemplo)
        ws[f"F{row}"] = num_nouns

        # ---------------- Métrica: Número de aplicações reivindicando prioridade ----------------
        priority_apps = []

        # Localiza a seção pelo ID appsClaimingPriority
        priority_section = soup.find('h3', id='appsClaimingPriority')
        if priority_section:
            # Pega a tabela imediatamente após o <h3>
            table = priority_section.find_next('div', class_='responsive-table style-scope patent-result')
            if table:
                rows = table.find_all('div', class_='tr style-scope patent-result')
                for row_div in rows:
                    # Ignora cabeçalhos (<th>)
                    if row_div.find('span', class_='th style-scope patent-result'):
                        continue

                    # Tenta pegar data-result do state-modifier
                    state = row_div.find('state-modifier')
                    if state and state.has_attr('data-result'):
                        priority_apps.append(state['data-result'])
                    else:
                        # Senão, pega o texto do primeiro <span> da linha
                        span = row_div.find('span', class_='td nowrap style-scope patent-result')
                        if span and span.get_text(strip=True):
                            priority_apps.append(span.get_text(strip=True))

        num_priority_apps = len(priority_apps)
        ws[f"G{row}"] = num_priority_apps

        # ---------------- Métrica: Frequência das 10 palavras mais comuns na descrição que aparecem no título (preciso com spaCy) ----------------

        # Localiza o elemento da descrição
        desc_element = driver.find_elements(By.XPATH,
                                            '/html/body/search-app/search-result/search-ui/div/div/div/div/div/'
                                            'result-container/patent-result/div/div/div/div[2]/div[1]/section/'
                                            'patent-text/div/section')

        if desc_element:
            description_text = desc_element[0].text.strip()
        else:
            description_text = ""

        # Processa a descrição com spaCy
        doc_desc = nlp(description_text)


        # Mantém apenas palavras alfabéticas, não stopwords
        desc_words = [token.lemma_.lower() for token in doc_desc if token.is_alpha and not token.is_stop]


        # Conta frequência das palavras
        word_freq = Counter(desc_words)


        # Pega as 10 mais comuns
        top_10_words = [w for w, _ in word_freq.most_common(10)]
        print(top_10_words)

        # Processa o título com spaCy
        doc_title = nlp(title_text)
        title_words = [token.lemma_.lower() for token in doc_title if token.is_alpha and not token.is_stop]

        # Soma quantas vezes as 10 palavras mais frequentes da descrição aparecem no título
        freq_in_title = sum(title_words.count(w) for w in top_10_words)

        # Salva no Excel (por exemplo, na coluna H)
        ws[f"H{row}"] = freq_in_title

        print(f"  Inventores: {num_inventors}, Imagens: {len(image_tags)}, Backward: {len(backward)}, "
              f"Forward: {len(cited_by)}, Substantivos: {num_nouns}, "
              f"Prioridade: {num_priority_apps}, Similaridade Título-Descrição: {ws[f'H{row}'].value}")

    except TimeoutException:
        print("  Tempo esgotado ao acessar a página")
        ws[f"H{row}"] = "ERRO"
    except Exception as e:
        print(f"  Erro inesperado: {e}")
        ws[f"H{row}"] = "ERRO"

# Fecha navegador e salva Excel
driver.quit()
wb.save(excel_path)
print("Processo concluído. Dados salvos no Excel.")