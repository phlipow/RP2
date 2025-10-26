from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import spacy
from collections import Counter
import re
from datetime import datetime

# Caminho do Excel
excel_path = r"C:\Users\Licitações\Downloads\WEBSCRAPPING.xlsx"

# Abre o Excel
wb = load_workbook(excel_path)
ws = wb["Planilha1"]

# Configurações do Chrome
chrome_options = Options()
# chrome_options.add_argument("--headless")  # descomente se quiser rodar sem abrir janela
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

# Carrega modelo spaCy
nlp = spacy.load("en_core_web_sm")

# Contadores de progresso
total_links = 0
links_modificados = 0
links_sem_dados = 0


def extract_from_soup(soup, driver):
    """Extrai todas as informações da patente e retorna em um dicionário."""
    result = {}

    # Inventores
    inventors = []
    important_people = soup.select("dl.important-people.style-scope.patent-result")
    if important_people:
        dl = important_people[0]
        dt_inventor = dl.find("dt", string=lambda x: x and "Inventor" in x)
        if dt_inventor:
            for dd in dt_inventor.find_all_next("dd"):
                state = dd.find("state-modifier")
                if state and state.has_attr("data-inventor"):
                    inventors.append(state["data-inventor"])
                else:
                    break
    result["inventors_count"] = len(inventors)

    # Imagens
    image_tags = soup.select("section#thumbnails img")
    result["images_count"] = len(image_tags)

    # Backward citations
    backward = []
    citations_section = soup.find("h3", id="patentCitations")
    if citations_section:
        table = citations_section.find_next("div", class_="responsive-table")
        if table:
            for m in table.find_all("state-modifier"):
                data = m.get("data-result", "")
                if "patent/" in data:
                    backward.append(data.split("patent/")[-1].split("/")[0])
    result["backward_count"] = len(backward)

    # Forward citations
    cited_by = []
    cited_by_section = soup.find("h3", id="citedBy")
    if cited_by_section:
        table = cited_by_section.find_next("div", class_="responsive-table")
        if table:
            for m in table.find_all("state-modifier"):
                data = m.get("data-result", "")
                if "patent/" in data:
                    cited_by.append(data.split("patent/")[-1].split("/")[0])
    result["forward_count"] = len(cited_by)

    # Título e substantivos
    title_tag = soup.find("h1", class_="scroll-target style-scope patent-result")
    title_text = title_tag.get_text(strip=True) if title_tag else ""
    result["title_text"] = title_text
    doc = nlp(title_text)
    result["num_nouns"] = sum(1 for token in doc if token.pos_ == "NOUN")

    # Priority apps
    priority_apps = []
    priority_section = soup.find("h3", id="appsClaimingPriority")
    if priority_section:
        table = priority_section.find_next("div", class_="responsive-table style-scope patent-result")
        if table:
            for row_div in table.find_all("div", class_="tr style-scope patent-result"):
                if row_div.find("span", class_="th style-scope patent-result"):
                    continue
                state = row_div.find("state-modifier")
                if state and state.has_attr("data-result"):
                    priority_apps.append(state["data-result"])
    result["priority_count"] = len(priority_apps)

    # Similaridade título/descrição
    try:
        desc_element = wait.until(
            EC.presence_of_element_located((By.XPATH, "//patent-text/div/section"))
        )
        description_text = desc_element.text.strip()
    except TimeoutException:
        description_text = ""

    doc_desc = nlp(description_text)
    desc_words = [t.lemma_.lower() for t in doc_desc if t.is_alpha and not t.is_stop]
    word_freq = Counter(desc_words)
    top_10 = [w for w, _ in word_freq.most_common(10)]

    doc_title = nlp(title_text)
    title_words = [t.lemma_.lower() for t in doc_title if t.is_alpha and not t.is_stop]
    result["similarity_score"] = sum(title_words.count(w) for w in top_10)

    # NPL citations
    num_npl = 0
    npl_section = soup.find("h3", id="nplCitations")
    if npl_section:
        match = re.search(r"\((\d+)\)", npl_section.get_text(strip=True))
        if match:
            num_npl = int(match.group(1))
        else:
            table = npl_section.find_next("div", class_="responsive-table")
            if table:
                num_npl = len(table.find_all("div", class_="tr style-scope patent-result"))
    result["npl_count"] = num_npl

    # Número de claims
    num_claims = 0
    for c in soup.find_all("div", class_="flex style-scope patent-result"):
        text = c.get_text(strip=True)
        if "Claims" in text or "Reivindicações" in text:
            span = c.find("span", class_="style-scope patent-result")
            if span:
                m = re.search(r"\((\d+)\)", span.get_text(strip=True))
                if m:
                    num_claims = int(m.group(1))
                    break
    result["num_claims"] = num_claims

    # Data de aplicação
    app_date_tag = soup.find("div", class_="filed style-scope application-timeline")
    if app_date_tag:
        app_date = app_date_tag.get_text(strip=True)
        try:
            app_date = datetime.strptime(app_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            pass
    else:
        app_date = ""
    result["app_date"] = app_date

    return result


# Loop principal
for row in range(2, 102):
    url_original = ws[f"A{row}"].value
    if not url_original:
        continue

    total_links += 1
    print(f"\n[Linha {row}] Acessando: {url_original}")
    urls_to_try = [url_original]

    # cria variação /BRPI se aplicável
    if "/patent/BR" in url_original and "BRPI" not in url_original:
        urls_to_try.append(url_original.replace("/patent/BR", "/patent/BRPI", 1))

    final_results = None
    link_used = "Link original"

    for idx, url in enumerate(urls_to_try, start=1):
        try:
            driver.get(url)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
            soup = BeautifulSoup(driver.page_source, "html.parser")
            results = extract_from_soup(soup, driver)

            found_any = any([
                results["inventors_count"],
                results["images_count"],
                results["backward_count"],
                results["forward_count"],
                results["num_nouns"],
                results["priority_count"],
                results["similarity_score"],
                results["npl_count"],
                results["num_claims"],
                bool(results["app_date"]),
                bool(results["title_text"]),
            ])

            if found_any:
                final_results = results
                if url != url_original:
                    ws[f"A{row}"] = url
                    link_used = "Link modificado"
                    links_modificados += 1
                print(f"  ✅ Dados encontrados (tentativa {idx})")
                break
            else:
                print(f"  ⚠️ Nenhum dado na tentativa {idx}")
        except TimeoutException:
            print(f"  ⏳ Timeout ao acessar {url}")
        except Exception as e:
            print(f"  ❌ Erro inesperado: {e}")

    # Grava resultados
    if final_results:
        ws[f"B{row}"] = final_results["inventors_count"]
        ws[f"C{row}"] = final_results["images_count"]
        ws[f"D{row}"] = final_results["backward_count"]
        ws[f"E{row}"] = final_results["forward_count"]
        ws[f"F{row}"] = final_results["num_nouns"]
        ws[f"G{row}"] = final_results["priority_count"]
        ws[f"H{row}"] = final_results["similarity_score"]
        ws[f"I{row}"] = final_results["npl_count"]
        ws[f"J{row}"] = final_results["num_claims"]
        ws[f"K{row}"] = final_results["app_date"]
        ws[f"L{row}"] = link_used
    else:
        ws[f"L{row}"] = "Sem dados"
        links_sem_dados += 1

# Finaliza
driver.quit()
wb.save(excel_path)

# 📊 Resumo final
print("\n====================== RESUMO ======================")
print(f"🔍 Total de patentes analisadas: {total_links}")
print(f"🧩 Links modificados:            {links_modificados}")
print(f"⚠️  Patentes sem dados:           {links_sem_dados}")
print("✅ Processo concluído. Dados salvos no Excel.")
print("====================================================")
